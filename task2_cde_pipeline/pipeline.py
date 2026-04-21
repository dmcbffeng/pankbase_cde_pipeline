#!/usr/bin/env python3
"""
PanKbase CDE Harmonization Pipeline

Reads raw consortia data (Excel) and a consortia-specific mapping config,
then extracts and transforms donor metadata into a clean, CDE-matched table.

Usage:
    python pipeline.py --data <data.xlsx> --mapping <mapping.json> --cde <cde_schema.json> --output <output.tsv>

Example:
    python pipeline.py \
        --data ../data/HPAP_Donor_Summary_197.xlsx \
        --mapping mappings/hpap_mapping.json \
        --cde ../task1_cde_definitions/pankbase_donor_cdes.json \
        --output output/hpap_cde_harmonized.tsv
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Transform functions
# ---------------------------------------------------------------------------

def transform_direct(value, _cfg):
    """Pass through as-is."""
    if value is None:
        return None
    return str(value).strip()


def transform_constant(_value, cfg):
    """Return a fixed constant value."""
    return cfg.get("constant_value")


def transform_numeric(value, _cfg):
    """Convert to numeric (float). Strip whitespace and non-numeric chars."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("none", "not available", "na", "n/a", ""):
        return None
    # Handle values like "<0.1" -- take the number
    s = re.sub(r'^[<>]=?\s*', '', s)
    # Remove trailing non-numeric (e.g., units accidentally included)
    s = re.sub(r'[^\d.\-]', '', s)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def transform_value_map(value, cfg):
    """Map source value to CDE permissible value using a lookup dict."""
    if value is None:
        return cfg.get("default", None)
    s = str(value).strip()
    if not s or s.lower() in ("none", "not available", "na", "n/a"):
        return cfg.get("default", None)
    vmap = cfg.get("value_map", {})
    if s in vmap:
        return vmap[s]
    # Case-insensitive fallback
    for k, v in vmap.items():
        if k.lower() == s.lower():
            return v
    return cfg.get("default", None)


def transform_binary_to_pos_neg(value, _cfg):
    """Convert 0/1 to Negative/Positive."""
    if value is None:
        return "Unknown"
    s = str(value).strip()
    if s == "1":
        return "Positive"
    elif s == "0":
        return "Negative"
    return "Unknown"


def transform_parse_value_cutoff(value, _cfg):
    """Parse 'value/cutoff' format (e.g., '119/20') and return the value part."""
    if value is None:
        return None
    s = str(value).strip()
    if "/" in s:
        parts = s.split("/")
        try:
            return float(parts[0])
        except ValueError:
            return None
    return transform_numeric(value, _cfg)


def transform_timedelta_to_hours(value, _cfg):
    """Convert timedelta-like string ('0 days HH:MM:SS') to decimal hours."""
    if value is None:
        return None
    s = str(value).strip()
    if s.lower() in ("nat", "none", ""):
        return None
    # Pattern: "X days HH:MM:SS"
    m = re.match(r'(\d+)\s*days?\s+(\d+):(\d+):(\d+)', s)
    if m:
        days = int(m.group(1))
        hours = int(m.group(2))
        minutes = int(m.group(3))
        seconds = int(m.group(4))
        total_hours = days * 24 + hours + minutes / 60 + seconds / 3600
        return round(total_hours, 2)
    # Try HH:MM:SS alone
    m = re.match(r'(\d+):(\d+):(\d+)', s)
    if m:
        hours = int(m.group(1))
        minutes = int(m.group(2))
        seconds = int(m.group(3))
        return round(hours + minutes / 60 + seconds / 3600, 2)
    return transform_numeric(s, _cfg)


def transform_minutes_to_hours(value, _cfg):
    """Convert minutes to hours."""
    num = transform_numeric(value, _cfg)
    if num is None:
        return None
    return round(num / 60, 2)


def transform_parse_duration_years(value, _cfg):
    """Parse text like '18 years' or '5 years' to numeric years."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("none", "na", "n/a", ""):
        return None
    m = re.search(r'([\d.]+)', s)
    if m:
        return float(m.group(1))
    return None


def transform_derive_hba1c_status(value, _cfg):
    """Derive diabetes status from HbA1c value."""
    num = transform_numeric(value, _cfg)
    if num is None:
        return "Unknown"
    if num < 5.7:
        return "Normal"
    elif num <= 6.4:
        return "Prediabetes"
    else:
        return "Diabetes"


def transform_derive_ethnicity_from_race(value, _cfg):
    """Derive ethnicity from HPAP race field (which sometimes contains 'Hispanic')."""
    if value is None:
        return "Unknown"
    s = str(value).strip().lower()
    if "hispanic" in s:
        return "Hispanic or Latino"
    return "Not Hispanic or Latino"


def transform_derive_donation_type(values, _cfg):
    """Derive donation type from dbd/dcd boolean fields."""
    if not isinstance(values, (list, tuple)):
        return "Unknown"
    dbd = str(values[0]).strip().lower() if values[0] else ""
    dcd = str(values[1]).strip().lower() if len(values) > 1 and values[1] else ""
    if dbd == "yes":
        return "DBD"
    elif dcd == "yes":
        return "DCD"
    return "Unknown"


def transform_datetime_diff_days(values, _cfg):
    """Calculate difference in days between two datetime values."""
    if not isinstance(values, (list, tuple)) or len(values) < 2:
        return None
    try:
        dt1 = values[0]
        dt2 = values[1]
        if dt1 is None or dt2 is None:
            return None
        if isinstance(dt1, str):
            dt1 = datetime.fromisoformat(dt1.replace(".000001", ""))
        if isinstance(dt2, str):
            dt2 = datetime.fromisoformat(dt2.replace(".000001", ""))
        diff = (dt2 - dt1).total_seconds() / 86400
        return round(diff, 1)
    except Exception:
        return None


def transform_presence_to_yes_no(value, _cfg):
    """Convert presence of a numeric value to Yes/No."""
    if value is None:
        return "No"
    try:
        float(str(value))
        return "Yes"
    except (ValueError, TypeError):
        return "No"


def transform_not_available(_value, _cfg):
    """Field not available in this source."""
    return None


TRANSFORM_REGISTRY = {
    "direct": transform_direct,
    "constant": transform_constant,
    "numeric": transform_numeric,
    "value_map": transform_value_map,
    "binary_to_pos_neg": transform_binary_to_pos_neg,
    "parse_value_cutoff": transform_parse_value_cutoff,
    "timedelta_to_hours": transform_timedelta_to_hours,
    "minutes_to_hours": transform_minutes_to_hours,
    "parse_duration_years": transform_parse_duration_years,
    "derive_hba1c_status": transform_derive_hba1c_status,
    "derive_ethnicity_from_race": transform_derive_ethnicity_from_race,
    "derive_donation_type": transform_derive_donation_type,
    "datetime_diff_days": transform_datetime_diff_days,
    "presence_to_yes_no": transform_presence_to_yes_no,
    "not_available": transform_not_available,
    "yes_no_to_value_list": transform_direct,
    "lookup_sheet": transform_not_available,  # handled specially
}


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_hpap_data(filepath, mapping):
    """Load HPAP-style data (named columns in a standard sheet)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = mapping.get("source_sheet", wb.sheetnames[0])
    ws = wb[sheet_name]

    # Read header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Read data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        rows.append(record)

    # Also load lookup sheets if needed
    lookup_sheets = {}
    for cde_name, cfg in mapping.get("column_mappings", {}).items():
        if cfg.get("transform") == "lookup_sheet":
            ls_name = cfg.get("lookup_sheet")
            if ls_name and ls_name in wb.sheetnames and ls_name not in lookup_sheets:
                ls_ws = wb[ls_name]
                ls_headers = [c.value for c in next(ls_ws.iter_rows(min_row=1, max_row=1))]
                ls_rows = []
                for r in ls_ws.iter_rows(min_row=2, values_only=True):
                    ls_rows.append(dict(zip(ls_headers, r)))
                lookup_sheets[ls_name] = ls_rows

    wb.close()
    return rows, lookup_sheets


def load_iidp_data(filepath, mapping):
    """Load IIDP-style data (positional columns with offset header row)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = mapping.get("source_sheet", wb.sheetnames[0])
    ws = wb[sheet_name]

    header_row = mapping.get("header_row", 8)
    data_start_row = mapping.get("data_start_row", 9)

    # Read headers
    headers = []
    for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row)):
        headers.append(cell.value)

    # Read data
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        # Skip empty rows
        if row[0] is None:
            continue
        rows.append(list(row))

    wb.close()
    return rows, {}


def get_source_value(record, source_column, mapping_cfg):
    """Extract source value(s) from a record based on column specification."""
    if source_column is None:
        return None

    if isinstance(source_column, list):
        # Multi-column source
        if isinstance(record, dict):
            return [record.get(col) for col in source_column]
        else:
            return [record[col] if col < len(record) else None for col in source_column]

    if isinstance(record, dict):
        return record.get(source_column)
    else:
        # Positional index
        if isinstance(source_column, int) and source_column < len(record):
            return record[source_column]
        return None


def do_lookup(donor_id, cfg, lookup_sheets):
    """Perform a lookup in an auxiliary sheet."""
    ls_name = cfg.get("lookup_sheet")
    if ls_name not in lookup_sheets:
        return None

    ls_rows = lookup_sheets[ls_name]
    key_col = cfg.get("lookup_key", "donor_ID")
    value_col = cfg.get("lookup_value", "value")
    filter_spec = cfg.get("lookup_filter")
    filter_exclude = cfg.get("lookup_filter_exclude")
    aggregate = cfg.get("lookup_aggregate")

    matches = []
    for row in ls_rows:
        if row.get(key_col) != donor_id:
            continue
        # Apply include filter
        if filter_spec:
            match = True
            for fk, fv in filter_spec.items():
                if isinstance(fv, list):
                    if row.get(fk) not in fv:
                        match = False
                elif row.get(fk) != fv:
                    match = False
            if not match:
                continue
        # Apply exclude filter
        if filter_exclude:
            skip = False
            for fk, fv in filter_exclude.items():
                if isinstance(fv, list):
                    if row.get(fk) in fv:
                        skip = True
                elif row.get(fk) == fv:
                    skip = True
            if skip:
                continue
        val = row.get(value_col)
        if val is not None:
            matches.append(str(val).strip())

    if not matches:
        return None
    if aggregate == "semicolon_join":
        return "; ".join(matches)
    return matches[0]


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_record(record, cde_schema):
    """Validate a harmonized record against the CDE schema. Returns warnings."""
    warnings = []
    cde_lookup = {c["cde_name"]: c for c in cde_schema.get("cdes", [])}

    for cde_name, value in record.items():
        if cde_name not in cde_lookup:
            continue
        cde = cde_lookup[cde_name]

        # Check required fields
        if cde.get("required") and value is None:
            warnings.append(f"Required field '{cde_name}' is missing")

        if value is None:
            continue

        # Check permissible values
        pv = cde.get("permissible_values")
        if pv and cde.get("data_type") == "value_list":
            if value not in pv:
                warnings.append(
                    f"Field '{cde_name}' value '{value}' not in permissible values: {pv}"
                )

        # Check numeric range
        if cde.get("data_type") == "number":
            try:
                num = float(value)
                mn = cde.get("min_value")
                mx = cde.get("max_value")
                if mn is not None and num < mn:
                    warnings.append(
                        f"Field '{cde_name}' value {num} below minimum {mn}"
                    )
                if mx is not None and num > mx:
                    warnings.append(
                        f"Field '{cde_name}' value {num} above maximum {mx}"
                    )
            except (ValueError, TypeError):
                warnings.append(
                    f"Field '{cde_name}' expected numeric but got '{value}'"
                )

        # Check pattern
        pattern = cde.get("pattern")
        if pattern and isinstance(value, str):
            if not re.match(pattern, value):
                warnings.append(
                    f"Field '{cde_name}' value '{value}' does not match pattern '{pattern}'"
                )

    return warnings


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run_pipeline(data_path, mapping_path, cde_path, output_path):
    """Execute the CDE harmonization pipeline."""
    # Load configs
    with open(mapping_path) as f:
        mapping = json.load(f)
    with open(cde_path) as f:
        cde_schema = json.load(f)

    consortium = mapping.get("consortium", "Unknown")
    print(f"[Pipeline] Consortium: {consortium}")
    print(f"[Pipeline] Loading data from: {data_path}")

    # Load source data
    if consortium == "IIDP":
        source_rows, lookup_sheets = load_iidp_data(data_path, mapping)
    else:
        source_rows, lookup_sheets = load_hpap_data(data_path, mapping)

    print(f"[Pipeline] Loaded {len(source_rows)} source records")

    # Get the ordered list of CDE field names
    cde_fields = [c["cde_name"] for c in cde_schema.get("cdes", [])]
    col_mappings = mapping.get("column_mappings", {})

    # Process each record
    harmonized = []
    all_warnings = []

    for i, record in enumerate(source_rows):
        # Get donor ID for lookups
        if isinstance(record, dict):
            donor_id_col = mapping.get("donor_id_column", "donor_ID")
            donor_id = record.get(donor_id_col)
        else:
            donor_id_col = mapping.get("donor_id_column", 0)
            donor_id = record[donor_id_col] if donor_id_col < len(record) else None

        output_record = {}

        for cde_name in cde_fields:
            cfg = col_mappings.get(cde_name, {})
            transform_name = cfg.get("transform", "not_available")
            source_col = cfg.get("source_column")

            # Handle lookup_sheet transform specially
            if transform_name == "lookup_sheet":
                value = do_lookup(donor_id, cfg, lookup_sheets)
                post = cfg.get("post_transform")
                if post and post in TRANSFORM_REGISTRY:
                    value = TRANSFORM_REGISTRY[post](value, cfg)
                output_record[cde_name] = value
                continue

            # Get source value
            raw_value = get_source_value(record, source_col, cfg)

            # Apply transform
            fn = TRANSFORM_REGISTRY.get(transform_name, transform_not_available)
            value = fn(raw_value, cfg)
            output_record[cde_name] = value

        # Validate
        warnings = validate_record(output_record, cde_schema)
        if warnings:
            for w in warnings:
                all_warnings.append(f"Record {i} ({donor_id}): {w}")

        harmonized.append(output_record)

    print(f"[Pipeline] Harmonized {len(harmonized)} records")
    print(f"[Pipeline] Validation warnings: {len(all_warnings)}")

    # Write output
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w") as f:
        # Header
        f.write("\t".join(cde_fields) + "\n")
        # Data
        for rec in harmonized:
            vals = []
            for field in cde_fields:
                v = rec.get(field)
                if v is None:
                    vals.append("")
                else:
                    # Escape tabs and newlines that would break TSV format
                    s = str(v).replace("\t", " ").replace("\r", " ").replace("\n", " ")
                    vals.append(s)
            f.write("\t".join(vals) + "\n")

    print(f"[Pipeline] Output written to: {output_path}")

    # Write validation report
    report_path = output_path.with_suffix(".validation_report.txt")
    with open(report_path, "w") as f:
        f.write(f"PanKbase CDE Harmonization Validation Report\n")
        f.write(f"Consortium: {consortium}\n")
        f.write(f"Source: {data_path}\n")
        f.write(f"Date: {datetime.now().isoformat()}\n")
        f.write(f"Total records: {len(harmonized)}\n")
        f.write(f"Total warnings: {len(all_warnings)}\n")
        f.write(f"\n{'='*60}\n\n")

        # Summary statistics
        f.write("Field Completeness:\n")
        for field in cde_fields:
            non_null = sum(1 for r in harmonized if r.get(field) is not None)
            pct = non_null / len(harmonized) * 100 if harmonized else 0
            f.write(f"  {field}: {non_null}/{len(harmonized)} ({pct:.1f}%)\n")

        f.write(f"\n{'='*60}\n\n")
        f.write("Warnings:\n")
        if all_warnings:
            for w in all_warnings[:200]:  # Cap at 200
                f.write(f"  {w}\n")
            if len(all_warnings) > 200:
                f.write(f"  ... and {len(all_warnings) - 200} more\n")
        else:
            f.write("  No warnings.\n")

    print(f"[Pipeline] Validation report written to: {report_path}")

    return harmonized, all_warnings


def main():
    parser = argparse.ArgumentParser(
        description="PanKbase CDE Harmonization Pipeline"
    )
    parser.add_argument(
        "--data", required=True, help="Path to source data file (Excel)"
    )
    parser.add_argument(
        "--mapping", required=True, help="Path to consortia mapping JSON"
    )
    parser.add_argument(
        "--cde",
        default=str(
            Path(__file__).parent.parent / "task1_cde_definitions" / "pankbase_donor_cdes.json"
        ),
        help="Path to CDE schema JSON",
    )
    parser.add_argument(
        "--output", required=True, help="Path for output TSV file"
    )
    args = parser.parse_args()

    run_pipeline(args.data, args.mapping, args.cde, args.output)


if __name__ == "__main__":
    main()
